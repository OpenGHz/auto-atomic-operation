"""Read the w2_b2 sheet from the panel-info Excel workbook and emit a
panel-assembly YAML that uses the **flat-list** placements format.

Usage:
    python examples/panel_assembly/w2_b2/generate_layout_from_excel.py \
        "C:/Users/10529/Downloads/点按工站面板信息记录.xlsx"

    # With explicit output path:
    python examples/panel_assembly/w2_b2/generate_layout_from_excel.py \
        "C:/Users/10529/Downloads/点按工站面板信息记录.xlsx" \
        --output examples/panel_assembly/w2_b2/layout_w2_b2.yaml
"""

from __future__ import annotations

import argparse
import math
import re
import warnings
from pathlib import Path
from typing import Any

import openpyxl

# ── Excel column indices (1-based) ──────────────────────────────────────
COL_TYPE = 2  # B: Switch / Knob / Stop
COL_SEQ = 3  # C: sequence number (1-37)
COL_QTY = 4  # D: quantity
COL_CLASS = 5  # E: class_name  e.g. Switch_Quantity1_06
COL_BTN_CLS = 6  # F: button_class
COL_COORD = 8  # H: distance-from-centre coordinate
COL_POS = 9  # I: (wall, row, col)  e.g. (2,4,1)
COL_NOTES = 11  # K: notes (size, diameter, …)

# ── Known y-values per wall-row (mm) ───────────────────────────────────
ROW_Y_MAP: dict[int, float] = {
    4: 120.0,
    5: 0.0,
    6: -120.0,
}

SHEET_NAME = "w2_b2"
DATA_START_ROW = 3  # first row with actual placement data


# ── coordinate parsing ──────────────────────────────────────────────────


def _parse_coord(raw: Any, *, fallback_y: float | None) -> tuple[float, float]:
    """Parse the H-column value into ``(x, y)`` in millimetres.

    Three representations exist in the workbook:
    * A string with a comma  – ``"350,120"`` / ``"-70,0"`` / ``"350,-120"``
    * A bare integer whose comma was eaten by Excel's thousands separator –
      ``350120`` (should be ``350,120``).
    * ``None`` – caller must handle.
    """
    if raw is None:
        raise ValueError("Coordinate cell is empty.")

    if isinstance(raw, str):
        parts = [p.strip() for p in raw.split(",")]
        if len(parts) == 2:
            return float(parts[0]), float(parts[1])
        raise ValueError(f"Cannot parse coordinate string: {raw!r}")

    # Numeric – likely "x,yyy" eaten by thousands-separator formatting.
    value = float(raw)
    if fallback_y is not None:
        # Recover x by treating the number as string-concatenation of x and y.
        y_str = str(int(abs(fallback_y)))  # e.g. "120"
        val_str = str(int(value))
        if val_str.endswith(y_str) and len(val_str) > len(y_str):
            x_str = val_str[: -len(y_str)]
            return float(x_str), fallback_y

    # Fallback: truncation-based recovery (works for y = ±120).
    x = int(value / 1000)
    y = value - x * 1000
    if fallback_y is not None and not math.isclose(y, fallback_y, abs_tol=1.0):
        warnings.warn(
            f"Recovered y={y} differs from expected y={fallback_y} for raw={raw!r}; "
            f"using fallback y.",
            stacklevel=2,
        )
        y = fallback_y
    return float(x), float(y)


def _parse_wall_pos(raw: Any) -> tuple[int, int, int] | None:
    """Parse ``(wall, row, col)`` from the I-column, e.g. ``'(2,4,1)'``."""
    if raw is None:
        return None
    text = str(raw).strip().strip("()")
    parts = [p.strip() for p in text.split(",")]
    if len(parts) != 3:
        return None
    return int(parts[0]), int(parts[1]), int(parts[2])


# ── main extraction ─────────────────────────────────────────────────────


def _extract_placements(
    excel_path: str | Path,
) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb[SHEET_NAME]

    placements: list[dict[str, Any]] = []
    seen_cols: dict[tuple[int, int], int] = {}  # (row, col) → seq for dup detection

    for excel_row in range(DATA_START_ROW, ws.max_row + 1):
        seq_raw = ws.cell(row=excel_row, column=COL_SEQ).value
        if seq_raw is None:
            continue
        seq = int(seq_raw)

        # ── type ──
        type_raw = ws.cell(row=excel_row, column=COL_TYPE).value
        if type_raw is None:
            # Auto-correct: rows 37-39 are missing type; assume Switch.
            type_name = "Switch"
            warnings.warn(
                f"seq {seq}: type is empty — auto-corrected to 'Switch'.",
                stacklevel=2,
            )
        else:
            type_name = str(type_raw).strip()

        class_name = ws.cell(row=excel_row, column=COL_CLASS).value
        if class_name is None:
            # Build a fallback class_name from type + seq.
            class_name = f"{type_name}_{seq}"
            warnings.warn(
                f"seq {seq}: class_name is empty — auto-generated '{class_name}'.",
                stacklevel=2,
            )
        else:
            class_name = str(class_name).strip()

        # ── wall position (row, col) ──
        wall_pos = _parse_wall_pos(ws.cell(row=excel_row, column=COL_POS).value)
        if wall_pos is None:
            warnings.warn(
                f"seq {seq}: wall position is empty — skipping.",
                stacklevel=2,
            )
            continue
        _, w_row, w_col = wall_pos
        fallback_y = ROW_Y_MAP.get(w_row)

        # ── duplicate (row, col) detection ──
        rc_key = (w_row, w_col)
        if rc_key in seen_cols:
            old_seq = seen_cols[rc_key]
            # Auto-correct: assign next available col.
            max_col_in_row = max(c for (r, c) in seen_cols if r == w_row)
            w_col = max_col_in_row + 1
            warnings.warn(
                f"seq {seq}: position ({w_row},{rc_key[1]}) already used by seq {old_seq} "
                f"— auto-corrected to ({w_row},{w_col}).",
                stacklevel=2,
            )
        seen_cols[(w_row, w_col)] = seq

        # ── coordinate ──
        coord_raw = ws.cell(row=excel_row, column=COL_COORD).value
        if coord_raw is None:
            warnings.warn(
                f"seq {seq}: coordinate is empty — skipping.",
                stacklevel=2,
            )
            continue
        x, y = _parse_coord(coord_raw, fallback_y=fallback_y)

        # ── notes ──
        notes = ws.cell(row=excel_row, column=COL_NOTES).value
        notes_str = str(notes).strip() if notes else None

        entry: dict[str, Any] = {
            "seq": seq,
            "type": type_name,
            "class_name": class_name,
            "xml": f"objects/{type_name}_{seq}.xml",
            "pos": [x, y],
            "row": w_row,
            "col": w_col,
        }
        if notes_str:
            entry["notes"] = notes_str

        placements.append(entry)

    return placements


# ── YAML generation ─────────────────────────────────────────────────────


def _format_yaml(placements: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    lines.append("# Auto-generated from Excel sheet 'w2_b2'")
    lines.append(
        "# Re-generate: python examples/panel_assembly/w2_b2/"
        "generate_layout_from_excel.py <excel_path>"
    )
    lines.append("")
    lines.append("panel:")
    lines.append("  xml: w2_b2_panel_base.xml")
    lines.append("  attach_to_body: panel_mount")
    lines.append("")
    lines.append("output_xml: generated/w2_b2_panel_assembly.xml")
    lines.append("")
    lines.append("layout:")
    lines.append("  coordinate_scale: 0.001    # mm -> m")
    lines.append("  face_axes: [x, z]")
    lines.append("  base_pos: [0.0, 30.0, 0.0]")
    lines.append("  default_quat: [0.7071068, -0.7071068, 0.0, 0.0]")
    lines.append("  remove_root_joints: true")
    lines.append("")
    lines.append("placements:")

    for p in placements:
        seq = p["seq"]
        typ = p["type"]
        row = p["row"]
        col = p["col"]
        notes = p.get("notes", "")
        comment = f"seq {seq} | {typ} | wall ({row},{col})"
        if notes:
            comment += f" | {notes}"
        lines.append(f"  # {comment}")
        lines.append(f"  - xml: {p['xml']}")
        lines.append(f"    pos: [{p['pos'][0]}, {p['pos'][1]}]")
        lines.append(f"    row: {row}")
        lines.append(f"    col: {col}")

    lines.append("")
    return "\n".join(lines)


# ── CLI ─────────────────────────────────────────────────────────────────


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate w2_b2 panel layout YAML from Excel workbook."
    )
    parser.add_argument("excel", type=Path, help="Path to the Excel workbook.")
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output YAML path. Defaults to layout_w2_b2.yaml next to this script.",
    )
    return parser


def main() -> None:
    args = _build_parser().parse_args()
    placements = _extract_placements(args.excel)

    output = args.output or (Path(__file__).parent / "layout_w2_b2.yaml")
    yaml_text = _format_yaml(placements)
    output.write_text(yaml_text, encoding="utf-8")

    print(f"Generated {output} with {len(placements)} placements.")
    for p in placements:
        print(
            f"  seq={p['seq']:>2}  type={p['type']:<6}  "
            f"pos=({p['pos'][0]:>7.1f}, {p['pos'][1]:>6.1f})  "
            f"row={p['row']}  col={p['col']}  xml={p['xml']}"
        )


if __name__ == "__main__":
    main()
